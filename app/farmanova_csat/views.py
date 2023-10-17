from django.shortcuts import render,redirect
from django.db import connection
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
import json
from openpyxl import Workbook, load_workbook
from django.db.models import Q

from apps.form.models import Form
from apps.division.models import Division
from apps.question.models import Question
from apps.salesPoint.models import SalesPoint
from apps.contact.models import Contact
from apps.response.models import Response
from apps.survey.models import Survey
from apps.zone.models import Zone


#DEF=============
def querySQL(SQL,Num=None):
    cursor = connection.cursor()
    cursor.execute(SQL)
    row = cursor.fetchall()
    if Num == 1:
        row= row[0][0]
    return row


#INDEX
def index(request):
    if not request.user.is_authenticated:
        return redirect('login')
    groups= []
    for group in request.user.groups.all():
        groups.append(group.name)

    if "salesPoint manager" in groups:
        salesPoints = SalesPoint.objects.filter(users=request.user)
        return render(request, 'indexSalesPoints.html', {'object': salesPoints[0]})

    elif "anonymous user" in groups:
        salesPoints = SalesPoint.objects.filter(users=request.user)
        return render(request, 'indexAnonymous.html', {'object_list': salesPoints})

    else:
        if "zone manager" in groups:
            zone = Zone.objects.filter(user=request.user)
            query = f"""
                SELECT d.name Division, z.name Zona, sp.name Sucursal, f.name Estandar, AVG(r.response) Calificacion, COUNT(DISTINCT s.bill_id) Cantidad
                  FROM [sicreo].[dbo].[survey_survey] s
                LEFT JOIN contact_contact c ON s.contact_id = c.id
                LEFT JOIN form_form f ON s.form_id = f.id
                LEFT JOIN salesPoint_salespoint sp ON s.sales_point_id = sp.id
                LEFT JOIN division_division d ON sp.division_id = d.id
                LEFT JOIN zone_zone z ON sp.zone_id = z.id
                RIGHT JOIN response_response r ON s.id = r.survey_id
                LEFT JOIN question_question q ON q.id = r.question_id
                WHERE z.id = { zone[0].id }
                GROUP BY d.name, z.name, sp.name, f.name
                ORDER BY Calificacion"""
            
            html_file = 'indexZone.html'

        else:
            query = """
                SELECT d.name Division, z.name Zona, sp.name Sucursal, f.name Estandar, AVG(r.response) Calificacion, COUNT(DISTINCT s.bill_id) Cantidad
                  FROM [sicreo].[dbo].[survey_survey] s
                LEFT JOIN contact_contact c ON s.contact_id = c.id
                LEFT JOIN form_form f ON s.form_id = f.id
                LEFT JOIN salesPoint_salespoint sp ON s.sales_point_id = sp.id
                LEFT JOIN division_division d ON sp.division_id = d.id
                LEFT JOIN zone_zone z ON sp.zone_id = z.id
                RIGHT JOIN response_response r ON s.id = r.survey_id
                LEFT JOIN question_question q ON q.id = r.question_id
                GROUP BY d.name, z.name, sp.name, f.name
                ORDER BY Calificacion"""
            
            html_file = 'index.html'

        data = querySQL(query)
        s = i = cr = eo = 0
        sQty = iQty = crQty = eoQty = 0
        sCount = iCount = crCount = eoCount = 0
        data_raw = {}
        for x in data:
            data_raw[x[2]] = {"s":[],"i":[],"cr":[],"eo":[], "zona":x[1], "cadena":x[0], "cantidad":0 }

        for x in data:
            row_data = list(x)
            if row_data[3] == 'Sorprender':
                data_raw[x[2]]["s"].append(row_data[4])
                sCount += 1
                s += row_data[4]
                sQty = row_data[5]
                #if row_data[5] > sQty: sQty = row_data[5]
            elif row_data[3] == 'Involúcrate':
                data_raw[x[2]]["i"].append(row_data[4])
                iCount += 1
                i += row_data[4]
                iQty += row_data[5]
                #if row_data[5] > iQty: iQty = row_data[5]
            elif row_data[3] == 'Confiabilidad y Respeto':
                data_raw[x[2]]["cr"].append(row_data[4])
                crCount += 1
                cr += row_data[4]
                crQty += row_data[5]
                #if row_data[5] > crQty: crQty = row_data[5]
            elif row_data[3] == 'Empatía y Ordenados':
                data_raw[x[2]]["eo"].append(row_data[4])
                eoCount += 1
                eo += row_data[4]
                eoQty += row_data[5]
                #if row_data[5] > eoQty: eoQty = row_data[5]
            data_raw[x[2]]["cantidad"] += row_data[5]
        table_data = []
        for k,v in data_raw.items():
            len_s = len(v["s"]) if len(v["s"])>0 else 1
            len_i = len(v["i"]) if len(v["i"])>0 else 1
            len_cr = len(v["cr"]) if len(v["cr"])>0 else 1
            len_eo = len(v["eo"]) if len(v["eo"])>0 else 1
            s_table = 0
            for x in v["s"]: s_table += x
            i_table = 0
            for x in v["i"]: i_table += x
            cr_table = 0
            for x in v["cr"]: cr_table += x
            eo_table = 0
            for x in v["eo"]: eo_table += x
            s_table = s_table/len_s
            i_table = i_table/len_i
            cr_table = cr_table/len_cr
            eo_table = eo_table/len_eo
            avg_total = (s_table+i_table+cr_table+eo_table) / 4
            table_data.append([k,s_table,i_table,cr_table,eo_table, avg_total,v["zona"],v["cadena"],v["cantidad"]])
        sCount = 1 if sCount==0 else sCount
        iCount = 1 if iCount==0 else iCount
        crCount = 1 if crCount==0 else crCount
        eoCount = 1 if eoCount==0 else eoCount
        s = s/sCount
        i = i/iCount
        cr = cr/crCount
        eo = eo/eoCount
        print(table_data)
        print({"s":s,"i":i,"cr":cr,"eo":eo})

        return render(request, html_file, {"s":s,"i":i,"cr":cr,"eo":eo, "sQty":sQty, "iQty":iQty, "crQty":crQty, "eoQty":eoQty, "table_data":table_data})


#API
def apiForms(request):
    company_slug = request.GET.get('company')
    if company_slug != "null" or company_slug is not None:
        company = Division.objects.get(slug=company_slug)
        queryset = Form.objects.filter(
            Q(division=company) | Q(division=None)
        )
        queryset = list(queryset.values())
        queryset = [{"id":x["id"],"name":x["name"],"description":x["description"]} for x in queryset]
        print(queryset)

        data = {"data": queryset}
        return JsonResponse(data)
    else:
        return JsonResponse({"data":[]})


def apiQuestions(request):
    form_id = request.GET.get('form_id')
    company_slug = request.GET.get('company')

    company = Division.objects.get(slug=company_slug)
    queryset = Question.objects.filter(
        Q(division=company) | Q(division=None),
        Q(form_id=form_id)
    )
    queryset = list(queryset.values())
    queryset = [{"id":x["id"],"question":x["question"],"qty_responses":x["qty_responses"]} for x in queryset]

    data = {"data": queryset}
    return JsonResponse(data)


def apiGetCompany(request):
    sales_point = request.GET.get('sales_point')
    queryset = SalesPoint.objects.get(slug=sales_point)
    company = queryset.division.slug
    logo = queryset.division.logo_white.url
    data = {"data": company, "logo": logo}
    return JsonResponse(data)


def apiGetContact(request):
    try:
        contact_id = request.GET.get('contact_id')
        queryset = Contact.objects.get(contact_id=contact_id)
        contact = {
            'id': queryset.contact_id,
            'first_name': queryset.first_name,
            'last_name':  queryset.last_name,
            'phone': queryset.phone,
            'mobile': queryset.mobile,
            'email': queryset.email
        }
        data = {"data": contact}
        return JsonResponse(data)
    except:
        return JsonResponse({'error': "Contacto no encontrado"}, status=404)

@csrf_exempt
@require_POST
def apiCreateSurvey(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        try:
            contact = False
            contact_new = False
            try:
                try: 
                    id_contact = data.get('id')
                    contact_created = Contact.objects.get(contact_id=id_contact)
                    contact = True
                except:
                    try: mobile = data.get('mobile')
                    except: mobile = None
                    try: phone = data.get('phone')
                    except: phone = None
                    try: email = data.get('email')
                    except: email = None
                    firstname = data.get('firstname')
                    lastname = data.get('lastname')
                    contact = True
                    contact_new = True

            except Exception as e:
                print(e)

            if contact_new:
                try: contact_created = Contact.objects.create(contact_id= id_contact, first_name= firstname, last_name= lastname, phone= phone, mobile= mobile, email=email)
                except: contact = False
            
            bill_id = data.get('bill_id')
            form_id = data.get('form_id')
            form_id = Form.objects.get(id=form_id)
            responses = data.get('responses')
            sales_point = data.get('sales_point')
            sales_point = SalesPoint.objects.get(slug=sales_point)
            if contact:
                survey = Survey.objects.create(sales_point=sales_point, contact=contact_created, form=form_id, bill_id=bill_id)
            else:
                survey = Survey.objects.create(sales_point=sales_point, form=form_id, bill_id=bill_id)

            for response_content in responses:
                question = Question.objects.get(id=response_content["id"])
                Response.objects.create(survey=survey, question=question, response=response_content["response"])                
            
            return JsonResponse({'message': 'Success'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
    else:
        return JsonResponse({'error': 'Esta vista solo admite peticiones POST.'}, status=405)
    

#Reports
def download_general_report(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.method == 'POST':
        from_date = request.POST.get('from')
        to_date = request.POST.get('to')
        query = """
            SELECT c.full_name, s.bill_id, s.created_at, d.name Division, z.name Zona, sp.name Sucursal, f.name Estandar, q.question Pregunta, r.response Calificacion
              FROM [sicreo].[dbo].[survey_survey] s
            LEFT JOIN contact_contact c ON s.contact_id = c.id
            LEFT JOIN form_form f ON s.form_id = f.id
            LEFT JOIN salesPoint_salespoint sp ON s.sales_point_id = sp.id
            LEFT JOIN division_division d ON sp.division_id = d.id
            LEFT JOIN zone_zone z ON sp.zone_id = z.id
            RIGHT JOIN response_response r ON s.id = r.survey_id
            LEFT JOIN question_question q ON q.id = r.question_id
        """
        if from_date != "" and to_date != "":
            query += f"WHERE s.created_at BETWEEN '{from_date}' AND '{to_date}'"
        elif from_date != "":
            query += f"WHERE s.created_at > '{from_date}'"
        elif to_date != "":
            query += f"WHERE s.created_at < '{to_date}'"
        data = querySQL(query)

        # Crea un nuevo libro de Excel
        wb = load_workbook(filename="./farmanova_csat/base.xlsx")
        ws = wb["Data"]
        buddy_list = []

        for row in data:
            buddy_list.append(list(row))

        for x in range(len(buddy_list)):
            ws[f"A{x+2}"] = buddy_list[x][0]
            ws[f"B{x+2}"] = buddy_list[x][1]
            ws[f"C{x+2}"] = buddy_list[x][2].date()
            ws[f"D{x+2}"] = buddy_list[x][2]
            ws[f"E{x+2}"] = buddy_list[x][3]
            ws[f"F{x+2}"] = buddy_list[x][4]
            ws[f"G{x+2}"] = buddy_list[x][5]
            ws[f"H{x+2}"] = buddy_list[x][6]
            ws[f"I{x+2}"] = buddy_list[x][7]
            ws[f"J{x+2}"] = buddy_list[x][8]

        # Define el nombre del archivo
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_sicreo.xlsx'

        # Guarda el libro de Excel en la respuesta
        wb.save(response)

        return response
    return HttpResponse("Acceso no permitido.")


def download_zone_report(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    zone = Zone.objects.filter(user=request.user)
    if request.method == 'POST':
        from_date = request.POST.get('from')
        to_date = request.POST.get('to')
        query = f"""
            SELECT c.full_name, s.bill_id, s.created_at, d.name Division, z.name Zona, sp.name Sucursal, f.name Estandar, q.question Pregunta, r.response Calificacion
              FROM [sicreo].[dbo].[survey_survey] s
            LEFT JOIN contact_contact c ON s.contact_id = c.id
            LEFT JOIN form_form f ON s.form_id = f.id
            LEFT JOIN salesPoint_salespoint sp ON s.sales_point_id = sp.id
            LEFT JOIN division_division d ON sp.division_id = d.id
            LEFT JOIN zone_zone z ON sp.zone_id = z.id
            RIGHT JOIN response_response r ON s.id = r.survey_id
            LEFT JOIN question_question q ON q.id = r.question_id
            WHERE z.id = { zone[0].id }
        """
        if from_date != "" and to_date != "":
            query += f"AND s.created_at BETWEEN '{from_date}' AND '{to_date}'"
        elif from_date != "":
            query += f"AND s.created_at > '{from_date}'"
        elif to_date != "":
            query += f"AND s.created_at < '{to_date}'"
        data = querySQL(query)

        # Crea un nuevo libro de Excel
        wb = load_workbook(filename="./farmanova_csat/base.xlsx")
        ws = wb["Data"]

        buddy_list = []

        for row in data:
            buddy_list.append(list(row))

        for x in range(len(buddy_list)):
            ws[f"A{x+2}"] = buddy_list[x][0]
            ws[f"B{x+2}"] = buddy_list[x][1]
            ws[f"C{x+2}"] = buddy_list[x][2].date()
            ws[f"D{x+2}"] = buddy_list[x][2]
            ws[f"E{x+2}"] = buddy_list[x][3]
            ws[f"F{x+2}"] = buddy_list[x][4]
            ws[f"G{x+2}"] = buddy_list[x][5]
            ws[f"H{x+2}"] = buddy_list[x][6]
            ws[f"I{x+2}"] = buddy_list[x][7]
            ws[f"J{x+2}"] = buddy_list[x][8]

        # Define el nombre del archivo
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_sicreo.xlsx'

        # Guarda el libro de Excel en la respuesta
        wb.save(response)

        return response
    return HttpResponse("Acceso no permitido.")
