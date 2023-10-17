from django.shortcuts import render,redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db import connection
from openpyxl import Workbook, load_workbook
from django.http import HttpResponse
import datetime

from .models import SalesPoint
from .forms import SalesPointForm
from apps.zone.models import Zone


#DEF=============
def querySQL(SQL,Num=None):
    cursor = connection.cursor()
    cursor.execute(SQL)
    row = cursor.fetchall()
    if Num == 1:
        row= row[0][0]
    return row

class SalesPointListView(ListView):
    template_name = 'salesPoint/salesPointList.html'
    queryset = SalesPoint.objects.all().order_by('id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        return context

def SalesPointListZoneView(request):
    if not request.user.is_authenticated:
        return redirect('login')

    zone = Zone.objects.filter(user=request.user)[0]
    sale_points = SalesPoint.objects.filter(zone=zone)

    print(sale_points)

    return render(request, 'salesPoint/salesPointListZone.html',{
        'object_list':sale_points
    })

class SalesPointDetailView(DetailView):
    model = SalesPoint
    template_name = 'salesPoint/salesPointDetails.html'

class SalesPointDetailViewZone(DetailView):
    model = SalesPoint
    template_name = 'salesPoint/salesPointDetailsAnonymous.html'

def SalesPointCreate(request):
    if not request.user.is_authenticated:
        return redirect('login')

    form = SalesPointForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        salesPoint = form.save()
        if salesPoint:
            return redirect('salesPointList') 

    return render(request, 'salesPoint/salesPointForm.html',{
        'form':form
    })

class SalesPointUpdate(UpdateView):
    model = SalesPoint
    form_class = SalesPointForm
    template_name = 'salesPoint/salesPointForm.html'
    success_url = reverse_lazy('salesPointList')    

class SalesPointDelete(DeleteView):
    model = SalesPoint
    template_name = 'salesPoint/salesPointDelete.html'
    success_url = reverse_lazy('salesPointList') 

#Reports
def download_sale_point_report(request, sales_point_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    query = f"""
        SELECT c.full_name, s.bill_id, s.created_at, d.name, sp.name, f.name, q.question, r.response
        FROM [fcsat].[dbo].[survey_survey] s
        LEFT JOIN contact_contact c ON s.contact_id = c.id
        LEFT JOIN form_form f ON s.form_id = f.id
        LEFT JOIN salesPoint_salespoint sp ON s.sales_point_id = sp.id
        LEFT JOIN division_division d ON sp.division_id = d.id
        RIGHT JOIN response_response r ON s.id = r.survey_id
        LEFT JOIN question_question q ON q.id = r.question_id
		WHERE sp.id = { sales_point_id }"""

    data = querySQL(query)
    # Crea un nuevo libro de Excel
    wb = load_workbook(filename="./apps/salesPoint/base.xlsx")
    ws = wb["Data"]
    buddy_list = []
    for row in data:
        buddy_list.append(list(row))

    for x in range(len(buddy_list)):
        print(buddy_list[x][0], type(buddy_list[x][0]))
        ws[f"A{x+2}"] = buddy_list[x][0] if buddy_list[x][0]!=" " else "Anonimo"
        ws[f"B{x+2}"] = buddy_list[x][1]
        ws[f"C{x+2}"] = buddy_list[x][2].date()
        ws[f"D{x+2}"] = buddy_list[x][2]
        ws[f"E{x+2}"] = buddy_list[x][3]
        ws[f"F{x+2}"] = buddy_list[x][4]
        ws[f"G{x+2}"] = buddy_list[x][5]
        ws[f"H{x+2}"] = buddy_list[x][6]
        ws[f"I{x+2}"] = buddy_list[x][7]

    # Define el nombre del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_sicreo_sucursal.xlsx'

    # Guarda el libro de Excel en la respuesta
    wb.save(response)

    return response
